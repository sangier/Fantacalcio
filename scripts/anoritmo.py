import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl

# Caricare i dati
data = pd.read_csv('../data/anoritmo/PlayerStats.csv')

def calc_mv_fv_historical(player_name, data):
    # Filtrare i dati del calciatore
    player_data = data[data['Nome'] == player_name]
    
    player_data = player_data[player_data['Pv'] >= 8]
    
    # Ordinare i dati in ordine decrescente di stagione
    player_data = player_data.sort_values('Season', ascending=False)
    
    # Sostituire i valori mancanti con 0
    player_data = player_data.fillna(0)
    
    # Calcolare MvH e FvH con pesi differenti
    if len(player_data) == 1:
        mvh = player_data['Mv'].iloc[0]
        fvh = player_data['Fm'].iloc[0]
    else:
        weights = [0.45 if season == '2022-23' else 0.55/(len(player_data['Season'].unique())-1) if (len(player_data['Season'].unique())-1) != 0 else 0.0 for season in player_data['Season']]
        mvh = (player_data['Mv'] * weights).sum() / sum(weights)
        fvh = (player_data['Fm'] * weights).sum() / sum(weights)
    
    return mvh, fvh

def classify_mv_fv_historical(row):
    mvh_classification = 'D'
    fvh_classification = 'D'
    
    # Verificare se mvh o fvh sono NaN e, in tal caso, assegnare la fascia 'New'
    if np.isnan(row['MvH']) or np.isnan(row['FvH']):
        return 'New', 'New'
    
    if row['R'] == 'A':  # Attaccanti
        if row['MvH'] >= 6.2:
            mvh_classification = 'A'
        elif 6.2 > row['MvH'] >= 5.9:
            mvh_classification = 'B'
        elif 5.9 > row['MvH']>= 5.7:
            mvh_classification = 'C'
        if row['FvH'] >= 7.2:
            fvh_classification = 'A'
        elif 7.2 > row['FvH'] >= 6.4:
            fvh_classification = 'B'
        elif 6.4 > row['FvH']>= 6:
            fvh_classification = 'C'
            
    elif row['R'] == 'C':  # Centrocampisti
        if row['MvH'] >= 6.2:
            mvh_classification = 'A'
        elif 6.2 > row['MvH'] >= 6:
            mvh_classification = 'B'
        elif 6 > row['MvH']>=5.8:
            mvh_classification = 'C'
        if row['FvH'] >= 6.8:
            fvh_classification = 'A'
        elif 6.8> row['FvH'] >= 6.2:
            fvh_classification = 'B'
        elif 6.2 > row['FvH']>= 6:
            fvh_classification = 'C'
            
    elif row['R'] == 'D':  # Difensori
        if row['MvH'] >= 6.15:
            mvh_classification = 'A'
        elif 6.15 > row['MvH'] >= 5.9:
            mvh_classification = 'B'
        elif 5.9 > row['MvH']>=5.7:
            mvh_classification = 'C'
        if row['FvH'] >= 6.35:
            fvh_classification = 'A'
        elif 6.35 > row['FvH'] >= 6:
            fvh_classification = 'B'
        elif 6 > row['FvH']>=5.8:
            fvh_classification = 'C'
            
    else:  # Portiere
        if row['MvH'] >= 6.1:
            mvh_classification = 'A'
        elif 6.1 > row['MvH'] >= 5.4:
            mvh_classification = 'B'
        elif 5.4 > row['MvH']>=5:
            mvh_classification = 'C'
        if row['FvH'] >= 5.4:
            fvh_classification = 'A'
        elif 5.4 > row['FvH'] >= 5:
            fvh_classification = 'B'
        elif 5 > row['FvH']>=4.8:
            fvh_classification = 'C'
            
    return mvh_classification, fvh_classification

# Ottenere tutti i nomi dei calciatori
player_names = data['Nome'].unique()

# Calcolare MvH e FvH per ogni calciatore e aggiungere al DataFrame
for player_name in player_names:
    mvh, fvh = calc_mv_fv_historical(player_name, data)
    data.loc[data['Nome'] == player_name, 'MvH'] = mvh
    data.loc[data['Nome'] == player_name, 'FvH'] = fvh

# Applicare la funzione di classificazione a ogni riga e impostare le colonne 'MvH_Classification' e 'FvH_Classification'
data['MvH_Classification'], data['FvH_Classification'] = zip(*data.apply(classify_mv_fv_historical, axis=1))

# Mantieni solo le righe corrispondenti alla stagione 2023-24
data = data[data['Season'] == '2023-24']

# Elimina tutti i duplicati basati sul nome del giocatore, mantenendo solo la prima occorrenza
data = data.drop_duplicates(subset='Nome', keep='first')

# Scrivere il DataFrame aggiornato in un nuovo file CSV
data.to_csv('../data/anoritmo/FinalPlayerStats.csv', index=False)
print(data)
# Ordinare i dati alfabeticamente per nome
data = data.sort_values('Nome')

# Scrivere il DataFrame in un file Excel
with pd.ExcelWriter('../data/anoritmo/FinalPlayerStats.xlsx') as writer:
    data.to_excel(writer, sheet_name='Tutti', index=False)
    data[data['R'] == 'P'].to_excel(writer, sheet_name='Portieri', index=False)
    data[data['R'] == 'D'].to_excel(writer, sheet_name='Difensori', index=False)
    data[data['R'] == 'C'].to_excel(writer, sheet_name='Centrocampisti', index=False)
    data[data['R'] == 'A'].to_excel(writer, sheet_name='Attaccanti', index=False)

# Aprire il file Excel e colorare le celle
wb = openpyxl.load_workbook('../data/anoritmo/FinalPlayerStats.xlsx')

# Definire i colori
greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
orangeFill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
lightBlueFill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

def color_cells(sheet):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        classification = row[-1].value  # modifica per puntare alla cella corretta
        fill = None
        if classification == 'A':
            fill = greenFill
        elif classification == 'B':
            fill = orangeFill
        elif classification == 'C':
            fill = yellowFill
        elif classification == 'New':
            fill = lightBlueFill
        
        if fill:
            for cell in row:
                cell.fill = fill

# Colorare le celle per ogni foglio
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    color_cells(sheet)

# Salvare il file
wb.save('../data/anoritmo/FinalPlayerStats_Color.xlsx')
